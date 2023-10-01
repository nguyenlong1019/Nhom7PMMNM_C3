import openpyxl
from openpyxl.chart import BarChart, Reference

# Tạo một tệp Excel mới
workbook = openpyxl.Workbook()

# Chọn sheet mặc định
sheet = workbook.active

# Đặt tiêu đề cho các cột
sheet['A1'] = 'Tổng số SV'
sheet['B1'] = 'Số SV A+'
sheet['C1'] = 'Số sinh viên F'
sheet['D1'] = 'Số điểm mà sinh viên đạt được nhiều nhất'

# Điền dữ liệu vào các ô
data = [
    (60, 32, 10, 9),
    (60, 30, 11, 8),
    (60, 25, 13, 8),
    (60, 22, 12, 7),
]

for row, row_data in enumerate(data, start=2):
    for col, value in enumerate(row_data, start=1):
        sheet.cell(row=row, column=col, value=value)

# Đặt chiều rộng của các cột sao cho dài bằng tiêu đề của cột
for col in sheet.iter_cols(min_col=1, max_col=4, min_row=1, max_row=1):
    max_length = 0
    column = col[0].column_letter  # Lấy tên cột (A, B, C, D)
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column].width = adjusted_width

# Tạo một biểu đồ cột
chart = BarChart()
chart.title = 'Biểu đồ điểm sinh viên'
chart.x_axis.title = 'Cột dữ liệu'
chart.y_axis.title = 'Giá trị'

# Tạo tham chiếu đến dữ liệu cột và thêm nó vào biểu đồ
data_ref = Reference(sheet, min_col=1, min_row=1, max_col=4, max_row=5)
category_ref = Reference(sheet, min_col=1, min_row=2, max_row=5)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(category_ref)

# Thêm biểu đồ vào tệp Excel và lùi sang bên phải 5 cột
sheet.add_chart(chart, "G5")

# Lưu tệp Excel
workbook.save('danh_sach_sinh_vien.xlsx')

# Đóng tệp Excel
workbook.close()
